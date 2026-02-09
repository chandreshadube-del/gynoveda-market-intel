"""
Expansion Intelligence Platform - Data Upload & Refresh
=========================================================
Multi-section upload page. Universal sections work for all verticals.
Healthcare-specific sections (NFHS-5, IVF data) appear conditionally.
"""

import streamlit as st
import pandas as pd
import os
from config import get_vertical, get_term, has_module
from db import save_table, load_table, get_db_status, push_all_to_neon, get_upload_log, is_db_mode

st.set_page_config(page_title="Upload & Refresh", page_icon="📤", layout="wide")

V = st.session_state.get("vertical", "healthcare")
vc = get_vertical(V)
pri = vc["color_primary"]
loc = get_term(V, "location")
locs = get_term(V, "locations")
txn = get_term(V, "transactions")
prod = get_term(V, "product")

st.markdown(f'<span style="background:linear-gradient(135deg,{pri},{vc["color_accent"]});color:white;padding:4px 14px;border-radius:20px;font-size:.8rem;font-weight:600">{vc["icon"]} {vc["name"]}</span>', unsafe_allow_html=True)
st.title(f"📤 Data Upload & Refresh")
st.markdown(f"Import your data to power analytics for **{vc['name']}**. All uploads are saved locally and synced to Neon if configured.")

# DB status
db = get_db_status()
st.info(f"{db['icon']} **Mode:** {db['mode']}")

st.divider()


# ==================================================================
# SECTION 1: Online Demand / Transaction Data
# ==================================================================
st.subheader(f"1️⃣ Online {txn} Data")
st.markdown(f"""
Upload your online {txn.lower()} data. Expected columns:
- **State/Region** - geographic identifier
- **Order quantity** - number of {txn.lower()}
- **Revenue** - transaction value (optional)
- **Year** - for trend analysis (optional)
- **City** - for city-level drill-down (optional)
- **Pincode** - for pincode-level analysis (optional)
""")

order_file = st.file_uploader(f"Upload {txn} data (CSV or Excel)", type=["csv", "xlsx", "xls"], key="orders")
if order_file:
    try:
        if order_file.name.endswith('.csv'):
            df = pd.read_csv(order_file)
        else:
            df = pd.read_excel(order_file)

        st.success(f"✅ Loaded {len(df):,} rows × {len(df.columns)} columns")
        st.dataframe(df.head(10), use_container_width=True)

        if st.button(f"💾 Save {txn} Data", key="save_orders"):
            # Auto-detect and save appropriate tables
            if 'state' in [c.lower() for c in df.columns]:
                state_col = next(c for c in df.columns if c.lower() == 'state')
                save_table("state_orders_summary", df)
                # Build master
                qty_col = next((c for c in df.columns if 'order' in c.lower() or 'qty' in c.lower() or 'quantity' in c.lower()), None)
                rev_col = next((c for c in df.columns if 'rev' in c.lower() or 'amount' in c.lower()), None)
                if qty_col:
                    master = df.groupby(state_col).agg(
                        total_orders=(qty_col, 'sum'),
                        **({f'total_revenue': (rev_col, 'sum')} if rev_col else {})
                    ).reset_index()
                    master.columns = ['state'] + [c for c in master.columns if c != state_col]
                    existing = load_table("master_state")
                    if not existing.empty and 'state' in existing.columns:
                        master = existing.merge(master, on='state', how='outer', suffixes=('', '_new'))
                        for c in ['total_orders', 'total_revenue']:
                            if f'{c}_new' in master.columns:
                                master[c] = master[f'{c}_new'].fillna(master.get(c, 0))
                                master.drop(f'{c}_new', axis=1, inplace=True)
                    save_table("master_state", master)

            if 'city' in [c.lower() for c in df.columns]:
                save_table("city_orders_summary", df)

            if 'year' in [c.lower() for c in df.columns]:
                save_table("year_trend", df.groupby('Year' if 'Year' in df.columns else 'year').agg(
                    total_orders=(next((c for c in df.columns if 'order' in c.lower() or 'qty' in c.lower()), df.select_dtypes('number').columns[0]), 'sum')
                ).reset_index())

            st.success("✅ Data saved successfully!")
            st.rerun()

    except Exception as e:
        st.error(f"Error reading file: {e}")

st.divider()


# ==================================================================
# SECTION 2: Location Performance Data
# ==================================================================
st.subheader(f"2️⃣ {loc} Performance Data")
st.markdown(f"""
Upload {loc.lower()} performance data. Expected columns:
- **{loc} name/ID**
- **Total {txn.lower()}** or quantity
- **First-time {get_term(V, 'customers').lower()}** (optional)
- **Pincode** (optional)
""")

loc_file = st.file_uploader(f"Upload {loc} data (CSV or Excel)", type=["csv", "xlsx", "xls"], key="locations")
if loc_file:
    try:
        if loc_file.name.endswith('.csv'):
            df_loc = pd.read_csv(loc_file)
        else:
            df_loc = pd.read_excel(loc_file)

        st.success(f"✅ Loaded {len(df_loc):,} rows × {len(df_loc.columns)} columns")
        st.dataframe(df_loc.head(10), use_container_width=True)

        if st.button(f"💾 Save {loc} Data", key="save_locations"):
            save_table("clinic_performance", df_loc)
            st.success("✅ Location data saved!")
            st.rerun()
    except Exception as e:
        st.error(f"Error: {e}")

st.divider()


# ==================================================================
# SECTION 3: Product / Category Data
# ==================================================================
if has_module(V, "product_intelligence"):
    st.subheader(f"3️⃣ {prod} / Category Data")
    st.markdown(f"Upload {prod.lower()}-level performance data with state breakdown.")

    prod_file = st.file_uploader(f"Upload {prod} data", type=["csv", "xlsx", "xls"], key="products")
    if prod_file:
        try:
            if prod_file.name.endswith('.csv'):
                df_prod = pd.read_csv(prod_file)
            else:
                df_prod = pd.read_excel(prod_file)

            st.success(f"✅ Loaded {len(df_prod):,} rows")
            st.dataframe(df_prod.head(10), use_container_width=True)

            if st.button(f"💾 Save {prod} Data", key="save_products"):
                save_table("product_state", df_prod)
                st.success("✅ Product data saved!")
                st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

    st.divider()


# ==================================================================
# SECTION 4: Pincode / Zip Data
# ==================================================================
st.subheader(f"4️⃣ Pincode / Catchment Data")
st.markdown(f"Upload pincode-level {loc.lower()} mapping for catchment analysis.")

pin_file = st.file_uploader("Upload pincode data", type=["csv", "xlsx", "xls"], key="pincodes")
if pin_file:
    try:
        if pin_file.name.endswith('.csv'):
            df_pin = pd.read_csv(pin_file)
        else:
            df_pin = pd.read_excel(pin_file)

        st.success(f"✅ Loaded {len(df_pin):,} rows")
        st.dataframe(df_pin.head(10), use_container_width=True)

        if st.button("💾 Save Pincode Data", key="save_pincodes"):
            save_table("pincode_clinic", df_pin)
            save_table("clinic_zip_summary", df_pin)
            st.success("✅ Pincode data saved!")
            st.rerun()
    except Exception as e:
        st.error(f"Error: {e}")

st.divider()


# ==================================================================
# SECTION 5: NFHS-5 Health Data (Healthcare only)
# ==================================================================
if V == "healthcare":
    st.subheader("5️⃣ NFHS-5 Health Indicators (Healthcare)")
    st.markdown("Upload NFHS-5 state or district data for health indicator analysis and CEI scoring.")

    nfhs_file = st.file_uploader("Upload NFHS-5 data", type=["csv", "xlsx", "xls"], key="nfhs")
    if nfhs_file:
        try:
            if nfhs_file.name.endswith('.csv'):
                df_nfhs = pd.read_csv(nfhs_file)
            else:
                df_nfhs = pd.read_excel(nfhs_file)

            st.success(f"✅ Loaded {len(df_nfhs):,} rows")
            st.dataframe(df_nfhs.head(10), use_container_width=True)

            level = st.radio("Data level", ["State", "District"], horizontal=True, key="nfhs_level")
            if st.button("💾 Save NFHS-5 Data", key="save_nfhs"):
                table = "nfhs5_state" if level == "State" else "nfhs5_district"
                save_table(table, df_nfhs)
                st.success(f"✅ NFHS-5 {level} data saved!")
                st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

    st.divider()


# ==================================================================
# SECTION 6: Infrastructure / Competition Data
# ==================================================================
st.subheader(f"6️⃣ Infrastructure & Competition Data")
st.markdown("Upload city-level infrastructure scores or competitor location data.")

infra_file = st.file_uploader("Upload infrastructure/competition data", type=["csv", "xlsx", "xls"], key="infra")
if infra_file:
    try:
        if infra_file.name.endswith('.csv'):
            df_infra = pd.read_csv(infra_file)
        else:
            df_infra = pd.read_excel(infra_file)

        st.success(f"✅ Loaded {len(df_infra):,} rows")
        st.dataframe(df_infra.head(10), use_container_width=True)

        save_as = st.radio("Save as", ["Infrastructure scores", "Competition map"], horizontal=True, key="infra_type")
        if st.button("💾 Save Data", key="save_infra"):
            table = "infra_city" if save_as == "Infrastructure scores" else "competition_map"
            save_table(table, df_infra)
            st.success(f"✅ {save_as} saved!")
            st.rerun()
    except Exception as e:
        st.error(f"Error: {e}")


# ==================================================================
# SECTION 7: Census District Demographics (All Verticals)
# ==================================================================
st.divider()
st.subheader("7️⃣ Census District Demographics")
st.markdown("""
Upload district-level Census 2011 data (640 districts). Accepts multi-sheet Excel files - select which sheets to import.

**Recognized datasets:** District demographics (population, literacy, workforce), Household infrastructure/assets, State-level health & spending indicators, Data sources reference.
""")

census_file = st.file_uploader("Upload Census data (Excel with multiple sheets)", type=["xlsx", "xls"], key="census")
if census_file:
    try:
        xls = pd.ExcelFile(census_file)
        sheet_names = xls.sheet_names
        st.info(f"📄 Found {len(sheet_names)} sheets: {', '.join(sheet_names)}")

        selected_sheets = st.multiselect("Select sheets to import:", sheet_names, default=sheet_names, key="census_sheets")

        if selected_sheets:
            preview_sheet = st.selectbox("Preview sheet:", selected_sheets, key="census_preview")
            df_preview = pd.read_excel(xls, sheet_name=preview_sheet)
            st.caption(f"**{preview_sheet}**: {len(df_preview):,} rows × {len(df_preview.columns)} columns")
            st.dataframe(df_preview.head(8), use_container_width=True)

            # Auto-detect table mapping
            sheet_table_map = {}
            for sn in selected_sheets:
                sn_lower = sn.lower().replace(" ", "_")
                if "demograph" in sn_lower or ("district" in sn_lower and "hh" not in sn_lower and "asset" not in sn_lower):
                    sheet_table_map[sn] = "census_district_demographics"
                elif "hh" in sn_lower or "infra" in sn_lower or "asset" in sn_lower:
                    sheet_table_map[sn] = "census_hh_assets"
                elif "health" in sn_lower or "spending" in sn_lower or "state" in sn_lower:
                    sheet_table_map[sn] = "state_health_spending"
                elif "source" in sn_lower:
                    sheet_table_map[sn] = None  # skip reference sheets
                else:
                    sheet_table_map[sn] = "census_district_demographics"

            st.markdown("**Auto-detected mapping:**")
            for sn, tbl in sheet_table_map.items():
                label = tbl.replace("_", " ").title() if tbl else "⏭️ Skip"
                st.caption(f"  {sn} → {label}")

            if st.button("💾 Save Census Data", key="save_census"):
                saved = 0
                for sn in selected_sheets:
                    tbl = sheet_table_map.get(sn)
                    if tbl:
                        df_sheet = pd.read_excel(xls, sheet_name=sn)
                        save_table(tbl, df_sheet)
                        saved += 1
                st.success(f"✅ Saved {saved} Census tables ({', '.join(t for t in sheet_table_map.values() if t)})")
                st.rerun()

    except Exception as e:
        st.error(f"Error reading Census file: {e}")


# ==================================================================
# SECTION 8: CEI Scoring Data
# ==================================================================
st.divider()
st.subheader("8️⃣ CEI Scoring Data (Pre-Computed)")
st.markdown("""
Upload pre-computed CEI district variables with scoring indices. This powers the **Market Scoring** page with real Census + NFHS-5 + HCES data instead of placeholders.

**Expected columns:** State, District, Population, Spending Index, Infrastructure Index, Health Need Index, Digital Index, Overall CEI Score.
""")

cei_file = st.file_uploader("Upload CEI Scoring data (Excel)", type=["xlsx", "xls"], key="cei")
if cei_file:
    try:
        xls_cei = pd.ExcelFile(cei_file)
        sheet_names_cei = xls_cei.sheet_names
        st.info(f"📄 Found {len(sheet_names_cei)} sheets: {', '.join(sheet_names_cei)}")

        selected_cei = st.multiselect("Select sheets to import:", sheet_names_cei, default=sheet_names_cei, key="cei_sheets")

        if selected_cei:
            preview_cei = st.selectbox("Preview sheet:", selected_cei, key="cei_preview")
            df_prev_cei = pd.read_excel(xls_cei, sheet_name=preview_cei)
            st.caption(f"**{preview_cei}**: {len(df_prev_cei):,} rows × {len(df_prev_cei.columns)} columns")
            st.dataframe(df_prev_cei.head(8), use_container_width=True)

            # Detect CEI columns
            if any("CEI" in str(c) or "Overall" in str(c) for c in df_prev_cei.columns):
                st.success("✅ CEI Score column detected - this will power the Market Scoring engine")

            # Auto-map sheets
            cei_table_map = {}
            for sn in selected_cei:
                sn_lower = sn.lower()
                if "variable" in sn_lower or "district" in sn_lower or "score" in sn_lower:
                    cei_table_map[sn] = "cei_district_scores"
                elif "method" in sn_lower:
                    cei_table_map[sn] = "cei_methodology"
                else:
                    cei_table_map[sn] = "cei_district_scores"

            if st.button("💾 Save CEI Data", key="save_cei"):
                saved_cei = 0
                for sn in selected_cei:
                    tbl = cei_table_map.get(sn)
                    if tbl:
                        df_sheet_cei = pd.read_excel(xls_cei, sheet_name=sn)
                        save_table(tbl, df_sheet_cei)
                        saved_cei += 1
                st.success(f"✅ Saved {saved_cei} CEI tables - Market Scoring page now powered by real data!")
                st.rerun()

    except Exception as e:
        st.error(f"Error reading CEI file: {e}")


# ==================================================================
# SECTION 9: City Classifications (Smart Cities / AMRUT)
# ==================================================================
st.divider()
st.subheader("9️⃣ City Classifications")
st.markdown("Upload Smart Cities, AMRUT, or tier classification data for market flagging.")

city_class_file = st.file_uploader("Upload city classification data (Excel)", type=["xlsx", "xls"], key="city_class")
if city_class_file:
    try:
        xls_cc = pd.ExcelFile(city_class_file)
        sheet_names_cc = xls_cc.sheet_names
        st.info(f"📄 Found {len(sheet_names_cc)} sheets: {', '.join(sheet_names_cc)}")

        selected_cc = st.multiselect("Select sheets:", sheet_names_cc, default=sheet_names_cc, key="cc_sheets")

        if selected_cc:
            preview_cc = st.selectbox("Preview:", selected_cc, key="cc_preview")
            df_prev_cc = pd.read_excel(xls_cc, sheet_name=preview_cc)
            st.caption(f"**{preview_cc}**: {len(df_prev_cc):,} rows × {len(df_prev_cc.columns)} columns")
            st.dataframe(df_prev_cc.head(8), use_container_width=True)

            # Auto-map
            cc_table_map = {}
            for sn in selected_cc:
                sn_lower = sn.lower()
                if "smart" in sn_lower:
                    cc_table_map[sn] = "smart_cities"
                elif "amrut" in sn_lower:
                    cc_table_map[sn] = "amrut_cities"
                else:
                    cc_table_map[sn] = None

            if st.button("💾 Save Classifications", key="save_cc"):
                saved_cc = 0
                for sn in selected_cc:
                    tbl = cc_table_map.get(sn)
                    if tbl:
                        df_cc = pd.read_excel(xls_cc, sheet_name=sn)
                        save_table(tbl, df_cc)
                        saved_cc += 1
                st.success(f"✅ Saved {saved_cc} classification tables!")
                st.rerun()

    except Exception as e:
        st.error(f"Error: {e}")


# ==================================================================
# SECTION 10: NTB Show Month Wise (Healthcare)
# ==================================================================
st.divider()
st.subheader("🔟 NTB Show Month Wise (Clinic Appointments)")
st.markdown("""
Upload the **NTB_Show_Month_Wise.xlsx** file. This contains clinic-level appointment counts
and show rates by month. The file has a special multi-row header that will be parsed automatically.

Expected structure: Area, City, Code, LaunchDate, Cabin, Region, then paired Appt/Show% columns per month.
""")

ntb_show_file = st.file_uploader("Upload NTB Show Month Wise (Excel)", type=["xlsx", "xls"], key="ntb_show")
if ntb_show_file:
    try:
        raw = pd.read_excel(ntb_show_file, header=None)
        st.caption(f"Raw shape: {raw.shape[0]} rows x {raw.shape[1]} cols")

        # Parse the multi-header structure
        # Row 0: label row (-, -, ..., Appt, Show%, Appt, Show%)
        # Row 1: actual headers (Area, City, Code, LaunchDate, Cabin, Region, All, All, month_dates...)
        header_labels = raw.iloc[0].tolist()
        header_dates = raw.iloc[1].tolist()

        # Extract month dates from row 1 (cols 8 onward are datetime month starts)
        months = []
        for i in range(8, len(header_dates), 2):
            val = header_dates[i]
            if hasattr(val, 'strftime'):
                months.append(val.strftime('%Y-%m'))
            elif pd.notna(val):
                months.append(str(val)[:7])

        st.caption(f"Detected {len(months)} months: {months[0]} to {months[-1]}" if months else "No months detected")

        # Parse data rows (row 2 onward)
        clinics = []
        summaries = []
        for idx in range(2, len(raw)):
            row = raw.iloc[idx]
            area = row[0]
            city = row[1]
            code = row[2]
            launch = row[3]
            cabins = row[4]
            region = row[5]
            total_appts = row[6]
            total_show = row[7]

            # Build monthly data
            monthly_data = {}
            for i, m in enumerate(months):
                col_a = 8 + i * 2
                col_s = 9 + i * 2
                if col_a < len(row):
                    appt_val = row[col_a]
                    show_val = row[col_s] if col_s < len(row) else None
                    monthly_data[f'appt_{m}'] = appt_val if pd.notna(appt_val) else None
                    monthly_data[f'show_{m}'] = show_val if pd.notna(show_val) else None

            record = {
                'Area': area if pd.notna(area) else None,
                'City': city if pd.notna(city) else None,
                'Code': code,
                'LaunchDate': launch,
                'Cabins': cabins,
                'Region': region if pd.notna(region) else None,
                'Total_Appts': total_appts,
                'Total_Show_Rate': total_show,
                **monthly_data
            }

            # Clinic rows have a numeric Code
            is_clinic = pd.notna(code) and str(code) != '-'
            try:
                int(float(str(code)))
                is_clinic = True
            except (ValueError, TypeError):
                is_clinic = False

            if is_clinic:
                clinics.append(record)
            else:
                summaries.append(record)

        df_clinics = pd.DataFrame(clinics)
        df_summaries = pd.DataFrame(summaries)

        st.success(f"Parsed **{len(df_clinics)} clinics** and **{len(df_summaries)} area/region summaries**")

        # Preview
        preview_tab1, preview_tab2 = st.tabs(["Clinic Data", "Area/Region Summaries"])
        with preview_tab1:
            st.dataframe(df_clinics.head(20), use_container_width=True, height=300, hide_index=True)
        with preview_tab2:
            st.dataframe(df_summaries.head(20), use_container_width=True, height=300, hide_index=True)

        if st.button("Save NTB Show Data", type="primary", key="save_ntb_show"):
            if not df_clinics.empty:
                save_table("ntb_show_clinic", df_clinics)

                # Also save to clinic_performance for Location Performance page
                perf = df_clinics[['Area', 'City', 'Code', 'LaunchDate', 'Cabins', 'Region',
                                   'Total_Appts', 'Total_Show_Rate']].copy()
                perf.columns = ['clinic_name', 'city_code', 'clinic_code', 'launch_date',
                                'cabins', 'region', 'total_appts', 'total_show_rate']
                # Add monthly columns
                appt_cols = [c for c in df_clinics.columns if c.startswith('appt_')]
                show_cols = [c for c in df_clinics.columns if c.startswith('show_')]
                for c in appt_cols + show_cols:
                    perf[c] = df_clinics[c].values
                save_table("clinic_performance", perf)

            if not df_summaries.empty:
                save_table("ntb_show_summary", df_summaries)

            st.success(f"Saved {len(df_clinics)} clinics to ntb_show_clinic + clinic_performance, {len(df_summaries)} summaries to ntb_show_summary")
            st.rerun()

    except Exception as e:
        st.error(f"Error parsing NTB Show file: {e}")


# ==================================================================
# SECTION 11: Clinic NTB ZipData (Healthcare)
# ==================================================================
st.divider()
st.subheader("1️⃣1️⃣ Clinic NTB ZipData (Transaction-Level)")
st.markdown(f"""
Upload the **ZipData_Clinic_NTB.xlsx** file. This contains transaction-level clinic data
(548K+ rows) which will be aggregated into clinic summaries and monthly trends.

Expected columns: Date, Quantity, Total, Customer Type, Clinic Loc, Region, Zone, Zip, SubCity, City, State, Launch Date, Cabin.

**Note:** Large files may take a minute to process.
""")

ntb_zip_file = st.file_uploader("Upload ZipData Clinic NTB (Excel)", type=["xlsx", "xls"], key="ntb_zip")
if ntb_zip_file:
    try:
        with st.spinner("Reading large Excel file... this may take a moment"):
            df_zip = pd.read_excel(ntb_zip_file)

        st.caption(f"Loaded: {df_zip.shape[0]:,} rows x {df_zip.shape[1]} cols")
        st.dataframe(df_zip.head(10), use_container_width=True, height=200, hide_index=True)

        # Standardize column names
        col_map = {}
        for c in df_zip.columns:
            cl = c.strip().lower().replace(' ', '_')
            col_map[c] = cl
        df_zip.rename(columns=col_map, inplace=True)

        # Detect key columns
        date_col = next((c for c in df_zip.columns if 'date' in c and 'launch' not in c), None)
        qty_col = next((c for c in df_zip.columns if 'quantity' in c or 'qty' in c), None)
        rev_col = next((c for c in df_zip.columns if 'total' in c and 'quantity' not in c), None)
        cust_col = next((c for c in df_zip.columns if 'customer' in c), None)
        clinic_col = next((c for c in df_zip.columns if 'clinic' in c), None)
        region_col = next((c for c in df_zip.columns if 'region' in c), None)
        zone_col = next((c for c in df_zip.columns if 'zone' in c), None)
        zip_col = next((c for c in df_zip.columns if 'zip' in c), None)
        city_col = next((c for c in df_zip.columns if c == 'city'), None)
        state_col = next((c for c in df_zip.columns if 'state' in c), None)
        cabin_col = next((c for c in df_zip.columns if 'cabin' in c), None)

        if date_col:
            df_zip[date_col] = pd.to_datetime(df_zip[date_col], errors='coerce')
            df_zip['month'] = df_zip[date_col].dt.to_period('M').astype(str)

        st.caption(f"Detected columns -> Date: {date_col}, Qty: {qty_col}, Revenue: {rev_col}, Clinic: {clinic_col}, Customer: {cust_col}")

        if st.button("Process & Save ZipData", type="primary", key="save_ntb_zip"):
            with st.spinner("Aggregating 548K+ rows..."):
                # 1. Clinic-level summary
                group_cols = [c for c in [clinic_col, region_col, zone_col, city_col, state_col] if c]
                if group_cols and qty_col:
                    agg_dict = {qty_col: 'sum'}
                    if rev_col:
                        agg_dict[rev_col] = 'sum'
                    if zip_col:
                        agg_dict[zip_col] = 'nunique'

                    clinic_summary = df_zip.groupby(group_cols).agg(agg_dict).reset_index()
                    # Rename columns
                    rename = {qty_col: 'total_qty'}
                    if rev_col:
                        rename[rev_col] = 'total_revenue'
                    if zip_col:
                        rename[zip_col] = 'unique_pincodes'
                    clinic_summary.rename(columns=rename, inplace=True)

                    # Add firsttime count
                    if cust_col:
                        ft_mask = df_zip[cust_col].str.lower().str.contains('first', na=False)
                        ft_agg = df_zip[ft_mask].groupby(group_cols)[qty_col].sum().reset_index()
                        ft_agg.rename(columns={qty_col: 'firsttime_qty'}, inplace=True)
                        clinic_summary = clinic_summary.merge(ft_agg, on=group_cols, how='left')
                        clinic_summary['firsttime_qty'] = clinic_summary['firsttime_qty'].fillna(0)

                    # Add transaction count
                    txn_count = df_zip.groupby(group_cols).size().reset_index(name='total_transactions')
                    clinic_summary = clinic_summary.merge(txn_count, on=group_cols, how='left')

                    save_table("ntb_zipdata_clinic", clinic_summary)

                    # Also save as clinic_zip_summary for Location Performance
                    save_table("clinic_zip_summary", clinic_summary)

                    st.success(f"Saved clinic summary: {len(clinic_summary)} locations")

                # 2. Monthly trend by clinic
                if clinic_col and 'month' in df_zip.columns and qty_col:
                    monthly_agg = {qty_col: 'sum'}
                    if rev_col:
                        monthly_agg[rev_col] = 'sum'
                    clinic_monthly = df_zip.groupby([clinic_col, 'month']).agg(monthly_agg).reset_index()
                    rename_m = {qty_col: 'qty'}
                    if rev_col:
                        rename_m[rev_col] = 'revenue'
                    clinic_monthly.rename(columns=rename_m, inplace=True)

                    # Add transaction count
                    txn_m = df_zip.groupby([clinic_col, 'month']).size().reset_index(name='transactions')
                    clinic_monthly = clinic_monthly.merge(txn_m, on=[clinic_col, 'month'], how='left')

                    save_table("ntb_zipdata_monthly", clinic_monthly)

                    # Also save as clinic_monthly_trend for Location Performance
                    save_table("clinic_monthly_trend", clinic_monthly)

                    st.success(f"Saved monthly trends: {len(clinic_monthly)} records across {clinic_monthly[clinic_col].nunique()} clinics")

                st.rerun()

    except Exception as e:
        st.error(f"Error processing ZipData: {e}")


# ==================================================================
# SECTION 12: 100-Clinic Expansion Strategy Workbook
# ==================================================================
st.divider()
st.subheader("1️⃣2️⃣ Expansion Strategy Workbook (All-in-One)")
st.markdown("""
Upload the **Gynoveda_100_Clinic_Expansion_Strategy_NEW.xlsx** workbook.
This single file contains 12 sheets that will be parsed automatically:
Revenue Projection (175 pincodes), Existing Clinics (61), Same-City Expansion,
New City Expansion, IVF Competitor Map, Web Order Demand, Implementation Roadmap,
Show% Analysis, Scenario Simulator, and Priority Tiers.
""")

exp_file = st.file_uploader("Upload Expansion Strategy Workbook (Excel)", type=["xlsx", "xls"], key="exp_strategy")
if exp_file:
    try:
        import openpyxl, json
        wb = openpyxl.load_workbook(exp_file, read_only=True, data_only=True)
        st.success(f"Loaded workbook with {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")

        if st.button("Parse & Save All 12 Sheets", type="primary", key="save_exp"):
            progress = st.progress(0, text="Parsing sheets...")
            saved = 0
            total_sheets = 12

            def parse_sheet(ws, min_row, max_row, header_row_data):
                """Generic sheet parser."""
                rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True))
                h = [str(c).strip().replace('\n', ' ') if c else f'col_{i}'
                     for i, c in enumerate(header_row_data)]
                df = pd.DataFrame(rows, columns=h[:len(rows[0])] if rows else h)
                return df

            # 1. Revenue Projection (175 pincodes)
            if "Revenue Projection" in wb.sheetnames:
                ws = wb["Revenue Projection"]
                rows = list(ws.iter_rows(min_row=19, max_row=193, values_only=True))
                cols = ['sno','pincode','area','city','type','parent_clinic','parent_avg_ntb',
                        'ntb_scale_pct','projected_ss_ntb','projected_show_pct','monthly_show_ntb',
                        'conv_pct','monthly_converted','sale_per_patient','ss_monthly_revenue',
                        'm1_rev','m2_rev','m3_rev','m4_rev','m5_rev','m6_rev',
                        'm7_rev','m8_rev','m9_rev','m10_rev','m11_rev','m12_rev',
                        'rev_12m_total','ntb_shows_12m','converted_12m','ntb_appts_12m']
                df = pd.DataFrame(rows, columns=cols).dropna(subset=['sno'])
                df['pincode'] = df['pincode'].astype(str).str.strip()
                save_table("revenue_projection_175", df)
                saved += 1
                st.caption(f"Revenue Projection: {len(df)} pincodes")

                # City rollup
                rows_c = list(ws.iter_rows(min_row=199, max_row=224, values_only=True))
                city_cols = ['city','pincodes','type','total_ss_ntb','avg_show_pct',
                             'total_ss_show_ntb','total_ss_monthly_rev','rev_12m_total',
                             'ntb_shows_12m','converted_12m']
                df_cr = pd.DataFrame(rows_c, columns=city_cols + [f'_x{i}' for i in range(21)])
                df_cr = df_cr[city_cols].dropna(subset=['city'])
                save_table("revenue_city_rollup", df_cr)

                # Assumptions
                assumptions = {
                    'conversion_pct': 0.40, 'sale_value': 18000,
                    'show_pct_adjustment': 0, 'new_city_default_show_pct': 0.20,
                    'same_city_ntb_scaling': 0.50, 'new_city_ntb_scaling': 1.0,
                    'ramp_m1': 0.30, 'ramp_m2': 0.50, 'ramp_m3': 0.65,
                    'ramp_m4': 0.80, 'ramp_m5': 0.90, 'ramp_m6': 1.0,
                }
                from db import DATA_DIR
                with open(os.path.join(DATA_DIR, "revenue_assumptions.json"), 'w') as f:
                    json.dump(assumptions, f, indent=2)
            progress.progress(1/total_sheets, text="Revenue Projection parsed...")

            # 2. Existing Clinics
            if "Existing Clinics (61)" in wb.sheetnames:
                ws2 = wb["Existing Clinics (61)"]
                rows2 = list(ws2.iter_rows(min_row=1, max_row=62, values_only=True))
                h2 = [str(c).strip().replace('\n',' ') if c else f'col_{i}' for i,c in enumerate(rows2[0])]
                df2 = pd.DataFrame(rows2[1:], columns=h2[:len(rows2[1])])
                save_table("existing_clinics_61", df2)
                saved += 1
                st.caption(f"Existing Clinics: {len(df2)} clinics")
            progress.progress(2/total_sheets)

            # 3. Same-City Expansion
            if "Same-City Expansion" in wb.sheetnames:
                ws3 = wb["Same-City Expansion"]
                rows3 = list(ws3.iter_rows(min_row=1, max_row=101, values_only=True))
                h3 = [str(c).strip() if c else f'col_{i}' for i,c in enumerate(rows3[0])]
                df3 = pd.DataFrame(rows3[1:], columns=h3[:len(rows3[1])])
                save_table("expansion_same_city", df3)
                saved += 1
            progress.progress(3/total_sheets)

            # 4. New City Expansion
            if "New City Expansion" in wb.sheetnames:
                ws4 = wb["New City Expansion"]
                rows4 = list(ws4.iter_rows(min_row=1, max_row=76, values_only=True))
                h4 = [str(c).strip() if c else f'col_{i}' for i,c in enumerate(rows4[0])]
                df4 = pd.DataFrame(rows4[1:], columns=h4[:len(rows4[1])])
                save_table("expansion_new_city", df4)
                saved += 1
            progress.progress(4/total_sheets)

            # 5. IVF Competitor Map
            if "IVF Competitor Map" in wb.sheetnames:
                ws5 = wb["IVF Competitor Map"]
                rows5 = list(ws5.iter_rows(min_row=1, max_row=16, values_only=True))
                h5 = [str(c).strip() if c else f'col_{i}' for i,c in enumerate(rows5[0])]
                df5 = pd.DataFrame(rows5[1:], columns=h5[:len(rows5[1])])
                save_table("ivf_competitor_map", df5)
                saved += 1
            progress.progress(5/total_sheets)

            # 6. Web Order Demand
            if "Web Order Demand" in wb.sheetnames:
                ws6 = wb["Web Order Demand"]
                rows6 = list(ws6.iter_rows(min_row=1, max_row=101, values_only=True))
                h6 = [str(c).strip() if c else f'col_{i}' for i,c in enumerate(rows6[0])]
                df6 = pd.DataFrame(rows6[1:], columns=h6[:len(rows6[1])])
                save_table("web_order_demand", df6)
                saved += 1
            progress.progress(6/total_sheets)

            # 7. Implementation Roadmap
            if "Implementation Roadmap" in wb.sheetnames:
                ws7 = wb["Implementation Roadmap"]
                rows7 = list(ws7.iter_rows(min_row=1, max_row=5, values_only=True))
                h7 = [str(c).strip() if c else f'col_{i}' for i,c in enumerate(rows7[0])]
                df7 = pd.DataFrame(rows7[1:], columns=h7[:len(rows7[1])])
                save_table("implementation_roadmap", df7)
                saved += 1
            progress.progress(7/total_sheets)

            # 8. Show% Analysis
            if "Show % of NTB Analysis" in wb.sheetnames:
                ws8 = wb["Show % of NTB Analysis"]
                rows8 = list(ws8.iter_rows(min_row=1, max_row=62, values_only=True))
                h8 = [str(c).strip().replace('\n',' ') if c else f'col_{i}' for i,c in enumerate(rows8[0])]
                df8 = pd.DataFrame(rows8[1:], columns=h8[:len(rows8[1])])
                save_table("show_pct_analysis", df8)
                saved += 1
            progress.progress(8/total_sheets)

            # 9. Scenario Simulator (clinic-level projections)
            if "Scenario Simulator" in wb.sheetnames:
                ws9 = wb["Scenario Simulator"]
                rows9 = list(ws9.iter_rows(min_row=26, max_row=87, values_only=True))
                h9 = [str(c).strip().replace('\n',' ') if c else f'col_{i}' for i,c in enumerate(rows9[0])]
                df9 = pd.DataFrame(rows9[1:], columns=h9[:len(rows9[1])])
                save_table("scenario_simulator_clinics", df9)
                saved += 1
            progress.progress(9/total_sheets)

            # 10. Show%-Based Expansion Plan (priority tiers)
            if "Show%-Based Expansion Plan" in wb.sheetnames:
                ws10 = wb["Show%-Based Expansion Plan"]
                # Priority tiers
                rows10a = list(ws10.iter_rows(min_row=30, max_row=34, values_only=True))
                h10a = [str(c).strip() if c else f'col_{i}' for i,c in enumerate(rows10a[0])]
                df10a = pd.DataFrame(rows10a[1:], columns=h10a[:len(rows10a[1])])
                save_table("expansion_priority_tiers", df10a)

                # Rank comparison
                rows10b = list(ws10.iter_rows(min_row=7, max_row=27, values_only=True))
                h10b = [str(c).strip() if c else f'col_{i}' for i,c in enumerate(rows10b[0])]
                df10b = pd.DataFrame(rows10b[1:], columns=h10b[:len(rows10b[1])])
                save_table("show_pct_rank_comparison", df10b)

                # Impact comparison
                rows10c = list(ws10.iter_rows(min_row=37, max_row=45, values_only=True))
                h10c = [str(c).strip() if c else f'col_{i}' for i,c in enumerate(rows10c[0])]
                df10c = pd.DataFrame(rows10c[1:], columns=h10c[:len(rows10c[1])])
                save_table("show_pct_impact_comparison", df10c)
                saved += 1
            progress.progress(10/total_sheets)

            wb.close()
            progress.progress(1.0, text="All sheets parsed!")
            st.success(f"Parsed and saved **{saved} sheet groups** from the Expansion Strategy workbook. "
                       "Navigate to **Revenue Simulator** and **Expansion Strategy** pages to explore.")
            st.rerun()

    except Exception as e:
        st.error(f"Error parsing Expansion Strategy workbook: {e}")
        import traceback
        st.code(traceback.format_exc())


# ==================================================================
# SYNC & LOG
# ==================================================================
st.divider()
st.subheader("🔄 Database Sync")

col_s1, col_s2 = st.columns(2)
with col_s1:
    if is_db_mode():
        if st.button("🔄 Push All to Neon", type="primary"):
            count = push_all_to_neon()
            st.success(f"✅ Synced {count} tables to Neon PostgreSQL.")
    else:
        st.info("Neon not configured. Data saved to local CSV files.")

with col_s2:
    log = get_upload_log()
    if log:
        st.markdown("**Recent Uploads:**")
        log_df = pd.DataFrame(log[-10:]).sort_values('ts', ascending=False)
        st.dataframe(log_df, use_container_width=True, height=200, hide_index=True)

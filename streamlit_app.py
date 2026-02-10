import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PAGE CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.set_page_config(
    page_title="Gynoveda North Star",
    page_icon="star",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CUSTOM CSS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600&display=swap');

/* Global */
.stApp { background: #0A0A0F; }
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; color: #E0E0E0; }
header[data-testid="stHeader"] { background: transparent; }
.block-container { padding-top: 1.5rem; padding-bottom: 1rem; max-width: 1400px; }

/* Hide streamlit branding */
#MainMenu, footer { visibility: hidden; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px; background: transparent; border-bottom: 1px solid rgba(255,255,255,0.04);
    padding-bottom: 0;
}
.stTabs [data-baseweb="tab"] {
    background: transparent; border-radius: 8px 8px 0 0; padding: 10px 24px;
    color: #666; font-weight: 600; font-size: 13px; border: none;
}
.stTabs [aria-selected="true"] {
    background: rgba(255,107,53,0.1); color: #FF6B35 !important;
    border-bottom: 2px solid #FF6B35;
}
.stTabs [data-baseweb="tab"]:hover { color: #ccc; }
.stTabs [data-baseweb="tab-panel"] { padding-top: 1.5rem; }

/* Metrics */
[data-testid="stMetric"] {
    background: linear-gradient(135deg, rgba(255,255,255,0.04) 0%, rgba(255,255,255,0.01) 100%);
    border: 1px solid rgba(255,255,255,0.06); border-radius: 14px;
    padding: 18px 20px;
}
[data-testid="stMetricLabel"] { color: #888 !important; font-size: 11px !important; letter-spacing: 0.06em; text-transform: uppercase; }
[data-testid="stMetricValue"] { color: #fff !important; font-size: 26px !important; font-weight: 700 !important; }
[data-testid="stMetricDelta"] { font-size: 12px !important; }

/* Selectbox / multiselect */
.stSelectbox > div > div, .stMultiSelect > div > div {
    background: rgba(255,255,255,0.04); border-color: rgba(255,255,255,0.08);
    color: #ccc; border-radius: 8px;
}

/* DataFrame */
.stDataFrame { border-radius: 12px; overflow: hidden; }

/* Divider */
hr { border-color: rgba(255,255,255,0.04) !important; }

/* KPI card custom */
.kpi-card {
    background: linear-gradient(135deg, rgba(255,255,255,0.04), rgba(255,255,255,0.01));
    border: 1px solid rgba(255,255,255,0.06); border-radius: 14px;
    padding: 20px 22px; position: relative; overflow: hidden;
}
.kpi-label { font-size: 11px; font-weight: 500; color: #888; letter-spacing: 0.07em; text-transform: uppercase; margin-bottom: 6px; }
.kpi-value { font-size: 28px; font-weight: 700; color: #fff; line-height: 1.1; }
.kpi-sub { font-size: 12px; font-weight: 500; margin-top: 6px; }

/* Section header */
.section-hdr { font-size: 18px; font-weight: 700; color: #fff; margin-bottom: 4px; letter-spacing: -0.01em; }
.section-sub { font-size: 12px; color: #555; margin-bottom: 16px; }

/* Zone card */
.zone-card {
    border-radius: 14px; padding: 20px; border: 1px solid rgba(255,255,255,0.06);
    background: rgba(255,255,255,0.02); transition: all 0.2s;
}
.zone-card:hover { border-color: rgba(255,255,255,0.12); }

/* Plotly chart background */
.js-plotly-plot .plotly .main-svg { background: transparent !important; }

/* Scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: rgba(255,255,255,0.1); border-radius: 3px; }

/* Expander */
.streamlit-expanderHeader { background: rgba(255,255,255,0.03); border-radius: 8px; }
</style>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA LOADING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
from pathlib import Path
DATA_DIR = Path(__file__).parent

@st.cache_data(ttl=3600)
def load_ntb_show():
    """Load NTB Show Month Wise - clinic appointments and show rates"""
    import openpyxl
    wb = openpyxl.load_workbook(DATA_DIR / "NTB_Show_Month_Wise.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Extract month headers
    months = []
    for i in range(8, len(rows[1]), 2):
        if rows[1][i] and hasattr(rows[1][i], "strftime"):
            months.append(rows[1][i].strftime("%Y-%m"))

    # Grand total row
    grand = rows[2]
    grand_total_appt = grand[6]
    grand_avg_show = grand[7]

    # Monthly grand totals
    monthly_grand = []
    for j, m in enumerate(months):
        col_appt = 8 + j * 2
        col_show = col_appt + 1
        appt = grand[col_appt] if col_appt < len(grand) else 0
        show = grand[col_show] if col_show < len(grand) else 0
        monthly_grand.append({"month": m, "appt": appt or 0, "show_pct": (show or 0) * 100})

    # Clinics
    clinics = []
    for r in rows[2:]:
        code = r[2]
        if code and code != "-":
            monthly = []
            for j, m in enumerate(months):
                col_a = 8 + j * 2
                col_s = col_a + 1
                a = r[col_a] if col_a < len(r) else 0
                s = r[col_s] if col_s < len(r) else 0
                monthly.append({"month": m, "appt": a or 0, "show_pct": (s or 0) * 100})
            launch_str = r[3].strftime("%Y-%m-%d") if hasattr(r[3], "strftime") else str(r[3])
            clinics.append({
                "area": r[0], "city": r[1], "code": str(r[2]),
                "launch": launch_str, "cabin": r[4], "region": r[5],
                "total_appt": r[6] or 0, "avg_show_pct": (r[7] or 0) * 100,
                "monthly": monthly,
            })

    # Region / sub-region summary rows
    summaries = []
    for r in rows[2:]:
        code = r[2]
        if (code == "-" or code is None) and r[0] not in ["Clinics", None]:
            summaries.append({
                "area": r[0], "total_appt": r[6] or 0,
                "avg_show_pct": (r[7] or 0) * 100,
            })

    return {
        "grand_total_appt": grand_total_appt,
        "grand_avg_show": grand_avg_show * 100,
        "monthly_grand": pd.DataFrame(monthly_grand),
        "clinics": clinics,
        "clinics_df": pd.DataFrame([{
            "Clinic": c["area"], "City": c["city"], "Code": c["code"],
            "Launch": c["launch"], "Cabin": c["cabin"], "Zone": c["region"],
            "Total Appts": c["total_appt"], "Avg Show %": round(c["avg_show_pct"], 1),
        } for c in clinics]),
        "summaries": summaries,
        "months": months,
    }


@st.cache_data(ttl=3600)
def load_clinic_firsttime():
    """Load Clinic-wise FirstTime by Pincode"""
    df = pd.read_excel(DATA_DIR / "Clinic_wise_FirstTime_TotalQuantity_by_Pincode.xlsx", sheet_name="FirstTime_by_Pincode")
    df.columns = ["Pincode", "Qty", "Revenue", "SubCity", "City", "State"]
    df["Pincode"] = df["Pincode"].astype(str)
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0).astype(int)
    df["Revenue"] = pd.to_numeric(df["Revenue"], errors="coerce").fillna(0)
    return df


@st.cache_data(ttl=3600)
def load_clinic_ntb():
    """Load ZipData_Clinic_NTB - transactional clinic data"""
    import openpyxl
    wb = openpyxl.load_workbook(DATA_DIR / "ZipData_Clinic_NTB.xlsx", read_only=True, data_only=True)
    ws = wb["Data"]

    from collections import defaultdict
    clinic_agg = defaultdict(lambda: {"qty": 0, "rev": 0, "region": "", "zone": "", "cabin": 0, "launch": "", "pincodes": set()})
    month_agg = defaultdict(lambda: {"qty": 0, "rev": 0, "visits": 0})
    zone_month = defaultdict(lambda: defaultdict(lambda: {"qty": 0, "rev": 0}))
    zone_agg = defaultdict(lambda: {"qty": 0, "rev": 0})

    for row in ws.iter_rows(min_row=2, values_only=True):
        date, qty, total, cust_type, clinic, region, zone = row[0], row[1] or 0, row[2] or 0, row[3], row[4] or "Unknown", row[5] or "", row[6] or ""
        zipcode, cabin, launch = row[7], row[12], row[11]

        clinic_agg[clinic]["qty"] += qty
        clinic_agg[clinic]["rev"] += total
        clinic_agg[clinic]["region"] = region
        clinic_agg[clinic]["zone"] = zone
        clinic_agg[clinic]["cabin"] = cabin or 0
        if launch and hasattr(launch, "strftime"):
            clinic_agg[clinic]["launch"] = launch.strftime("%Y-%m-%d")
        if zipcode:
            clinic_agg[clinic]["pincodes"].add(str(zipcode))

        if hasattr(date, "strftime"):
            ym = date.strftime("%Y-%m")
            month_agg[ym]["qty"] += qty
            month_agg[ym]["rev"] += total
            month_agg[ym]["visits"] += 1
            zone_month[zone][ym]["qty"] += qty
            zone_month[zone][ym]["rev"] += total

        if zone:
            zone_agg[zone]["qty"] += qty
            zone_agg[zone]["rev"] += total

    wb.close()

    clinic_df = pd.DataFrame([{
        "Clinic": k, "Zone": v["zone"], "Region": v["region"], "Cabin": v["cabin"],
        "Launch": v["launch"], "Qty": v["qty"], "Revenue": v["rev"],
        "Pincodes": len(v["pincodes"]),
        "Rev_L": round(v["rev"] / 100000, 1),
    } for k, v in clinic_agg.items()]).sort_values("Revenue", ascending=False)

    month_df = pd.DataFrame([{"month": k, **v} for k, v in month_agg.items()]).sort_values("month")
    month_df["rev_lakhs"] = month_df["rev"] / 100000

    zone_df = pd.DataFrame([{"Zone": k, "Qty": v["qty"], "Revenue": v["rev"], "Rev_Cr": round(v["rev"] / 10000000, 2)} for k, v in zone_agg.items() if k]).sort_values("Revenue", ascending=False)

    # Zone monthly for trend
    zone_month_records = []
    for z, months in zone_month.items():
        if z:
            for m, vals in months.items():
                zone_month_records.append({"Zone": z, "month": m, "qty": vals["qty"], "rev": vals["rev"]})
    zone_month_df = pd.DataFrame(zone_month_records).sort_values("month") if zone_month_records else pd.DataFrame()

    return {
        "clinic_df": clinic_df,
        "month_df": month_df,
        "zone_df": zone_df,
        "zone_month_df": zone_month_df,
    }


@st.cache_data(ttl=3600)
def load_ecom():
    """Load 1CX e-commerce orders"""
    import openpyxl
    wb = openpyxl.load_workbook(DATA_DIR / "1cx_order_qty_pincode_of_website__2020__2025__Curative__others.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]

    from collections import defaultdict
    year_agg = defaultdict(lambda: {"orders": 0, "qty": 0, "rev": 0})
    state_agg = defaultdict(lambda: {"orders": 0, "qty": 0, "rev": 0})
    city_agg = defaultdict(lambda: {"orders": 0, "qty": 0, "rev": 0, "state": ""})
    product_agg = defaultdict(lambda: {"orders": 0, "rev": 0})
    total_orders = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_orders += 1
        date, qty, total = row[0], row[3] or 0, row[4] or 0
        state, city = row[10] or "Unknown", row[9] or "Unknown"
        product_type = row[11] or "-"
        year = date.year if hasattr(date, "year") else "Unknown"

        year_agg[year]["orders"] += 1
        year_agg[year]["qty"] += qty
        year_agg[year]["rev"] += total
        state_agg[state]["orders"] += 1
        state_agg[state]["qty"] += qty
        state_agg[state]["rev"] += total
        city_agg[city]["orders"] += 1
        city_agg[city]["qty"] += qty
        city_agg[city]["rev"] += total
        city_agg[city]["state"] = state
        product_agg[product_type]["orders"] += 1
        product_agg[product_type]["rev"] += total

    wb.close()

    total_rev = sum(v["rev"] for v in year_agg.values())
    total_qty = sum(v["qty"] for v in year_agg.values())

    year_df = pd.DataFrame([{"Year": str(k), "Orders": v["orders"], "Qty": v["qty"], "Revenue": v["rev"], "Rev_Cr": round(v["rev"]/1e7, 1)} for k, v in year_agg.items()]).sort_values("Year")
    state_df = pd.DataFrame([{"State": k, "Orders": v["orders"], "Qty": v["qty"], "Revenue": v["rev"], "Rev_Cr": round(v["rev"]/1e7, 2)} for k, v in state_agg.items()]).sort_values("Revenue", ascending=False)
    city_df = pd.DataFrame([{"City": k, "State": v["state"], "Orders": v["orders"], "Qty": v["qty"], "Revenue": v["rev"], "Rev_Cr": round(v["rev"]/1e7, 2)} for k, v in city_agg.items()]).sort_values("Revenue", ascending=False)
    product_df = pd.DataFrame([{"Product": k, "Orders": v["orders"], "Revenue": v["rev"], "Rev_Cr": round(v["rev"]/1e7, 1)} for k, v in product_agg.items()]).sort_values("Revenue", ascending=False)

    return {
        "total_orders": total_orders, "total_rev": total_rev, "total_qty": total_qty,
        "year_df": year_df, "state_df": state_df, "city_df": city_df, "product_df": product_df,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# LOAD DATA
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ntb = load_ntb_show()
clinic_ft = load_clinic_firsttime()
clinic_ntb = load_clinic_ntb()
ecom = load_ecom()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ZONE_COLORS = {
    "West 1": "#FF6B35", "North 1": "#1E96FC", "West 2": "#F0C808",
    "South": "#2EC4B6", "East": "#9B5DE5", "North 2": "#F15BB5",
}

PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="DM Sans, sans-serif", color="#999", size=11),
    margin=dict(l=40, r=20, t=40, b=40),
    xaxis=dict(gridcolor="rgba(255,255,255,0.04)", zerolinecolor="rgba(255,255,255,0.04)"),
    yaxis=dict(gridcolor="rgba(255,255,255,0.04)", zerolinecolor="rgba(255,255,255,0.04)"),
    legend=dict(font=dict(size=11, color="#888"), bgcolor="rgba(0,0,0,0)"),
    hoverlabel=dict(bgcolor="#1a1a2e", font_size=12, font_family="DM Sans"),
)

def fmt_cr(val):
    return f"{val/1e7:.1f}" if val >= 1e7 else f"{val/1e5:.0f}L"

def fmt_lakhs(val):
    return f"{val/1e5:.1f}L"

def kpi_html(label, value, sub="", accent="#FF6B35"):
    return f"""
    <div class="kpi-card">
        <div style="position:absolute;top:-15px;right:-15px;width:70px;height:70px;border-radius:50%;background:{accent};opacity:0.05;filter:blur(18px);"></div>
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        <div class="kpi-sub" style="color:{accent};">{sub}</div>
    </div>
    """

def section_header(title, subtitle=""):
    st.markdown(f'<div class="section-hdr">{title}</div>', unsafe_allow_html=True)
    if subtitle:
        st.markdown(f'<div class="section-sub">{subtitle}</div>', unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HEADER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
hdr_left, hdr_right = st.columns([4, 1])
with hdr_left:
    st.markdown("""
    <div style="margin-bottom:8px;">
        <span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:#2EC4B6;box-shadow:0 0 10px #2EC4B6;margin-right:8px;vertical-align:middle;"></span>
        <span style="font-size:11px;color:#2EC4B6;font-weight:600;letter-spacing:0.1em;text-transform:uppercase;">Live &bull; North Star</span>
    </div>
    <h1 style="font-size:32px;font-weight:800;margin:0;letter-spacing:-0.03em;
        background:linear-gradient(135deg,#fff 40%,#FF6B35);
        -webkit-background-clip:text;-webkit-text-fill-color:transparent;">
        Gynoveda Expansion Command
    </h1>
    <p style="font-size:13px;color:#444;margin:6px 0 0;">
        61 Clinics &middot; 7,415 Pincodes &middot; 6 Zones &middot; Jan 2025 &rarr; Jan 2026
    </p>
    """, unsafe_allow_html=True)
with hdr_right:
    latest = clinic_ntb["month_df"].iloc[-1] if len(clinic_ntb["month_df"]) > 1 else None
    prev = clinic_ntb["month_df"].iloc[-2] if len(clinic_ntb["month_df"]) > 2 else None
    if latest is not None and prev is not None:
        mom = (latest["rev_lakhs"] - prev["rev_lakhs"]) / prev["rev_lakhs"] * 100
        arrow = "down_arrow" if mom < 0 else "up_arrow"
        st.markdown(f"""
        <div style="text-align:right;padding-top:16px;">
            <div style="font-size:11px;color:#555;">Latest Period</div>
            <div style="font-size:18px;font-weight:700;color:#fff;">{latest['month']}</div>
            <div style="font-size:13px;color:{'#F15BB5' if mom < 0 else '#2EC4B6'};font-weight:600;">
                {'&#x25BC;' if mom < 0 else '&#x25B2;'} {abs(mom):.1f}% MoM
            </div>
        </div>
        """, unsafe_allow_html=True)

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TABS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
tab1, tab2, tab3, tab4 = st.tabs([
    "Command Center",
    "Clinic Deep Dive",
    "Online to Offline Flywheel",
    "Zone Intelligence",
])

# ================================================================
# TAB 1: COMMAND CENTER
# ================================================================
with tab1:
    # KPI Row
    total_ntb_rev = clinic_ntb["month_df"]["rev"].sum()
    total_ntb_rev_cr = total_ntb_rev / 1e7

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    with k1:
        st.markdown(kpi_html("Total NTB Appointments", f"{ntb['grand_total_appt']:,.0f}", "Jan 25 - Jan 26", "#FF6B35"), unsafe_allow_html=True)
    with k2:
        st.markdown(kpi_html("Avg Show Rate", f"{ntb['grand_avg_show']:.1f}%", "National Average", "#2EC4B6"), unsafe_allow_html=True)
    with k3:
        st.markdown(kpi_html("Clinic NTB Revenue", f"Rs.{total_ntb_rev_cr:.0f}Cr", "Cumulative 13 months", "#F0C808"), unsafe_allow_html=True)
    with k4:
        st.markdown(kpi_html("Active Clinics", "61", "Across 6 Zones", "#9B5DE5"), unsafe_allow_html=True)
    with k5:
        st.markdown(kpi_html("Pincode Reach", "7,415", "Clinic Customers", "#1E96FC"), unsafe_allow_html=True)
    with k6:
        st.markdown(kpi_html("E-Com 1CX Orders", f"{ecom['total_orders']:,.0f}", f"Rs.{ecom['total_rev']/1e7:.1f}Cr Lifetime", "#F15BB5"), unsafe_allow_html=True)

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

    # Revenue Trajectory
    section_header("NTB Revenue Trajectory", "Monthly Clinic Revenue (Rs. Lakhs) & Show Rate %")

    mdf = clinic_ntb["month_df"].copy()
    ntb_monthly = ntb["monthly_grand"]

    fig_rev = make_subplots(specs=[[{"secondary_y": True}]])
    fig_rev.add_trace(
        go.Scatter(x=mdf["month"], y=mdf["rev_lakhs"], name="Revenue (Rs.L)",
                   fill="tozeroy", fillcolor="rgba(255,107,53,0.08)",
                   line=dict(color="#FF6B35", width=2.5), mode="lines"),
        secondary_y=False,
    )
    fig_rev.add_trace(
        go.Scatter(x=ntb_monthly["month"], y=ntb_monthly["show_pct"], name="Show Rate %",
                   line=dict(color="#2EC4B6", width=2, dash="dot"),
                   mode="lines+markers", marker=dict(size=5, color="#2EC4B6")),
        secondary_y=True,
    )
    fig_rev.update_layout(**PLOTLY_LAYOUT, height=340, title="")
    fig_rev.update_yaxes(title_text="Revenue (Rs. Lakhs)", secondary_y=False, gridcolor="rgba(255,255,255,0.04)", title_font=dict(size=11, color="#555"))
    fig_rev.update_yaxes(title_text="Show Rate %", secondary_y=True, gridcolor="rgba(255,255,255,0.02)", range=[10, 30], title_font=dict(size=11, color="#555"))
    st.plotly_chart(fig_rev, use_container_width=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # Zone Performance Matrix
    section_header("Zone Performance Matrix", "Revenue, Appointments & Show Rate by Zone")

    zdf = clinic_ntb["zone_df"].copy()
    # Merge with NTB show data for appt and show%
    zone_show_map = {}
    for s in ntb["summaries"]:
        if s["area"] in ZONE_COLORS:
            zone_show_map[s["area"]] = s

    zone_cards_data = []
    for _, row in zdf.iterrows():
        z = row["Zone"]
        show_info = zone_show_map.get(z, {})
        zone_cards_data.append({
            "zone": z, "rev_cr": row["Rev_Cr"], "qty": row["Qty"],
            "appt": show_info.get("total_appt", 0),
            "show_pct": show_info.get("avg_show_pct", 0),
            "color": ZONE_COLORS.get(z, "#666"),
        })

    cols = st.columns(3)
    for i, zc in enumerate(zone_cards_data):
        with cols[i % 3]:
            show_color = "#2EC4B6" if zc["show_pct"] >= 25 else "#F0C808" if zc["show_pct"] >= 20 else "#F15BB5"
            rev_pct = (zc["rev_cr"] / max(z["rev_cr"] for z in zone_cards_data)) * 100 if zone_cards_data else 0
            st.markdown(f"""
            <div class="zone-card" style="border-left:3px solid {zc['color']};margin-bottom:12px;">
                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
                    <span style="font-size:16px;font-weight:700;color:{zc['color']};">{zc['zone']}</span>
                </div>
                <div style="display:flex;gap:20px;">
                    <div><div style="font-size:10px;color:#555;">Revenue</div><div style="font-size:18px;font-weight:700;color:#fff;">Rs.{zc['rev_cr']}Cr</div></div>
                    <div><div style="font-size:10px;color:#555;">NTB Appts</div><div style="font-size:18px;font-weight:700;color:#fff;">{zc['appt']:,}</div></div>
                    <div><div style="font-size:10px;color:#555;">Show %</div><div style="font-size:18px;font-weight:700;color:{show_color};">{zc['show_pct']:.1f}%</div></div>
                </div>
                <div style="margin-top:10px;background:rgba(255,255,255,0.04);border-radius:4px;height:4px;overflow:hidden;">
                    <div style="width:{rev_pct:.0f}%;height:100%;background:{zc['color']};border-radius:4px;"></div>
                </div>
            </div>
            """, unsafe_allow_html=True)


# ================================================================
# TAB 2: CLINIC DEEP DIVE
# ================================================================
with tab2:
    cdf = clinic_ntb["clinic_df"].copy()

    # Merge with NTB show data
    ntb_clinics_df = ntb["clinics_df"].copy()
    cdf = cdf.merge(
        ntb_clinics_df[["Clinic", "Total Appts", "Avg Show %"]],
        on="Clinic", how="left"
    )
    cdf["Total Appts"] = cdf["Total Appts"].fillna(0).astype(int)
    cdf["Avg Show %"] = cdf["Avg Show %"].fillna(0)

    # Filters
    fcol1, fcol2, fcol3 = st.columns([2, 2, 3])
    with fcol1:
        sort_by = st.selectbox("Sort by", ["Revenue", "Qty", "Avg Show %", "Pincodes", "Total Appts"], index=0)
    with fcol2:
        zones_available = sorted(cdf["Zone"].dropna().unique().tolist())
        zone_filter = st.multiselect("Filter Zone", zones_available, default=[])
    with fcol3:
        cabin_filter = st.multiselect("Filter Cabin", sorted(cdf["Cabin"].dropna().unique().tolist()), default=[])

    display_df = cdf.copy()
    if zone_filter:
        display_df = display_df[display_df["Zone"].isin(zone_filter)]
    if cabin_filter:
        display_df = display_df[display_df["Cabin"].isin(cabin_filter)]

    sort_col = {"Revenue": "Revenue", "Qty": "Qty", "Avg Show %": "Avg Show %", "Pincodes": "Pincodes", "Total Appts": "Total Appts"}[sort_by]
    display_df = display_df.sort_values(sort_col, ascending=False).head(25)

    # Display table
    show_cols = ["Clinic", "Zone", "Region", "Cabin", "Launch", "Rev_L", "Qty", "Total Appts", "Avg Show %", "Pincodes"]
    st.dataframe(
        display_df[show_cols].rename(columns={"Rev_L": "Revenue (Rs.L)"}),
        use_container_width=True,
        hide_index=True,
        height=500,
        column_config={
            "Revenue (Rs.L)": st.column_config.ProgressColumn(
                "Revenue (Rs.L)", format="%.0f", min_value=0,
                max_value=float(display_df["Rev_L"].max()) if len(display_df) > 0 else 1,
            ),
            "Avg Show %": st.column_config.NumberColumn("Show %", format="%.1f%%"),
        }
    )

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    # Charts row
    ch1, ch2 = st.columns(2)
    with ch1:
        section_header("Top Clinics by Revenue", "Revenue in Rs. Lakhs, colored by Zone")
        top10 = display_df.head(10)
        fig_bar = go.Figure()
        for _, row in top10.iterrows():
            color = ZONE_COLORS.get(row["Zone"], "#666")
            fig_bar.add_trace(go.Bar(
                x=[row["Clinic"]], y=[row["Rev_L"]],
                marker_color=color, name=row["Zone"],
                showlegend=False, hovertemplate=f"{row['Clinic']}<br>Rs.{row['Rev_L']:.0f}L<br>Zone: {row['Zone']}<extra></extra>",
            ))
        fig_bar.update_layout(**PLOTLY_LAYOUT, height=320, showlegend=False)
        fig_bar.update_xaxes(tickangle=-35)
        st.plotly_chart(fig_bar, use_container_width=True)

    with ch2:
        section_header("Show Rate Radar by Region", "NTB Show % across metro regions")
        radar_data = [s for s in ntb["summaries"] if s["area"] in ["MMR","PMR","Gujarat","NCR","BMR","KMR","HMR","UP","Central","Punjab"]]
        if radar_data:
            fig_radar = go.Figure()
            fig_radar.add_trace(go.Scatterpolar(
                r=[r["avg_show_pct"] for r in radar_data],
                theta=[r["area"] for r in radar_data],
                fill="toself", fillcolor="rgba(255,107,53,0.1)",
                line=dict(color="#FF6B35", width=2),
                name="Show %",
            ))
            fig_radar.update_layout(
                **{k: v for k, v in PLOTLY_LAYOUT.items() if k not in ["xaxis", "yaxis"]},
                height=320,
                polar=dict(
                    bgcolor="rgba(0,0,0,0)",
                    radialaxis=dict(visible=True, range=[0, 35], gridcolor="rgba(255,255,255,0.06)", color="#555"),
                    angularaxis=dict(gridcolor="rgba(255,255,255,0.06)", color="#999"),
                ),
            )
            st.plotly_chart(fig_radar, use_container_width=True)

    # Clinic monthly trend selector
    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
    section_header("Individual Clinic Trajectory", "Select clinics to compare monthly appointment trends")

    clinic_names = [c["area"] for c in ntb["clinics"]]
    selected_clinics = st.multiselect("Select Clinics", clinic_names, default=clinic_names[:3])

    if selected_clinics:
        fig_clinic_trend = go.Figure()
        colors = ["#FF6B35", "#2EC4B6", "#9B5DE5", "#F0C808", "#1E96FC", "#F15BB5"]
        for i, cname in enumerate(selected_clinics):
            clinic_data = next((c for c in ntb["clinics"] if c["area"] == cname), None)
            if clinic_data:
                months = [m["month"] for m in clinic_data["monthly"]]
                appts = [m["appt"] for m in clinic_data["monthly"]]
                fig_clinic_trend.add_trace(go.Scatter(
                    x=months, y=appts, name=cname, mode="lines+markers",
                    line=dict(color=colors[i % len(colors)], width=2),
                    marker=dict(size=4),
                ))
        fig_clinic_trend.update_layout(**PLOTLY_LAYOUT, height=320, title="")
        fig_clinic_trend.update_yaxes(title_text="NTB Appointments")
        st.plotly_chart(fig_clinic_trend, use_container_width=True)


# ================================================================
# TAB 3: ONLINE TO OFFLINE FLYWHEEL
# ================================================================
with tab3:
    # Flywheel KPIs
    curative_rev = ecom["product_df"][ecom["product_df"]["Product"] == "Curative"]["Revenue"].sum()
    clinic_total_rev = clinic_ft["Revenue"].sum()

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(kpi_html("E-Com Lifetime Orders", f"{ecom['total_orders']:,}", f"Rs.{ecom['total_rev']/1e7:.1f}Cr Revenue", "#F15BB5"), unsafe_allow_html=True)
    with k2:
        st.markdown(kpi_html("Curative Share", f"{curative_rev/ecom['total_rev']*100:.0f}%", f"Rs.{curative_rev/1e7:.1f}Cr of Rs.{ecom['total_rev']/1e7:.1f}Cr", "#2EC4B6"), unsafe_allow_html=True)
    with k3:
        st.markdown(kpi_html("Clinic NTB Reach", f"{len(clinic_ft):,}", f"Pincodes | Rs.{clinic_total_rev/1e7:.0f}Cr Rev", "#FF6B35"), unsafe_allow_html=True)
    with k4:
        st.markdown(kpi_html("Total Clinic NTB Qty", f"{clinic_ft['Qty'].sum():,}", "First-time customers", "#F0C808"), unsafe_allow_html=True)

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

    # E-Com Yearly Trend
    ec1, ec2 = st.columns([2, 1])
    with ec1:
        section_header("E-Commerce 1CX Revenue Journey", "First-time customer orders by year (Rs. Crores)")
        fig_ecom = go.Figure()
        year_df = ecom["year_df"]
        fig_ecom.add_trace(go.Bar(
            x=year_df["Year"], y=year_df["Rev_Cr"], name="Revenue",
            marker_color=["#FF6B3555", "#FF6B3577", "#FF6B35", "#FF6B35", "#FF6B3577", "#FF6B3555"],
            text=year_df["Rev_Cr"].apply(lambda x: f"Rs.{x}Cr"),
            textposition="outside", textfont=dict(size=11, color="#ccc"),
        ))
        fig_ecom.update_layout(**PLOTLY_LAYOUT, height=300, showlegend=False)
        fig_ecom.update_yaxes(title_text="Revenue (Rs. Cr)")
        st.plotly_chart(fig_ecom, use_container_width=True)

    with ec2:
        section_header("Product Mix", "E-Com revenue by product type")
        fig_pie = go.Figure()
        pdf = ecom["product_df"]
        fig_pie.add_trace(go.Pie(
            labels=pdf["Product"], values=pdf["Revenue"],
            marker=dict(colors=["#FF6B35", "#2EC4B6", "#F0C808"]),
            textinfo="label+percent", textfont=dict(size=11),
            hole=0.5,
        ))
        fig_pie.update_layout(
            **{k: v for k, v in PLOTLY_LAYOUT.items() if k not in ["xaxis", "yaxis"]},
            height=300, showlegend=False,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    # City-Level Flywheel
    section_header("City-Level Online to Offline Flywheel", "E-commerce demand (blue) vs Clinic First-Time Revenue (orange) by city")

    # Build city comparison
    ecom_cities = ecom["city_df"].head(15).copy()
    clinic_city_agg = clinic_ft.groupby("City").agg({"Qty": "sum", "Revenue": "sum"}).reset_index()
    clinic_city_agg["Clinic_Rev_Cr"] = clinic_city_agg["Revenue"] / 1e7

    # Manual mapping for top cities
    city_map = {
        "Mumbai": "Mumbai", "Delhi": "Delhi", "Bengaluru": "Bengaluru",
        "Pune": "Pune", "Hyderabad": "Hyderabad", "Gurgaon": "Gurgaon",
        "Gautam Buddha Nagar": "Noida/GBN", "Ghaziabad": "Ghaziabad",
        "Ahmedabad": "Ahmedabad", "Lucknow": "Lucknow", "Kolkata": "Kolkata",
        "Goa": "Goa", "Kamrup": "Guwahati", "Nagpur": "Nagpur",
    }

    flywheel_data = []
    for _, row in ecom_cities.iterrows():
        city = row["City"]
        label = city_map.get(city, city)
        clinic_match = clinic_city_agg[clinic_city_agg["City"].str.contains(city, case=False, na=False)]
        clinic_rev = clinic_match["Clinic_Rev_Cr"].sum() if len(clinic_match) > 0 else 0
        flywheel_data.append({
            "City": label, "E-Com Rev (Cr)": row["Rev_Cr"], "Clinic Rev (Cr)": round(clinic_rev, 2),
        })

    fly_df = pd.DataFrame(flywheel_data)

    fig_fly = go.Figure()
    fig_fly.add_trace(go.Bar(
        y=fly_df["City"], x=fly_df["E-Com Rev (Cr)"], name="E-Com Rev (Rs.Cr)",
        orientation="h", marker_color="#1E96FC", opacity=0.7,
    ))
    fig_fly.add_trace(go.Bar(
        y=fly_df["City"], x=fly_df["Clinic Rev (Cr)"], name="Clinic Rev (Rs.Cr)",
        orientation="h", marker_color="#FF6B35", opacity=0.85,
    ))
    fig_fly.update_layout(**PLOTLY_LAYOUT, height=420, barmode="group", title="")
    fig_fly.update_xaxes(title_text="Revenue (Rs. Crores)")
    fig_fly.update_yaxes(autorange="reversed")
    st.plotly_chart(fig_fly, use_container_width=True)

    st.info("**Expansion Signal:** Gurgaon (Rs.1.19Cr e-com) and Ghaziabad (Rs.1.04Cr e-com) show strong online demand but have no dedicated clinic presence. These are prime candidates for new clinic placement.")

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    # State Conversion Table
    section_header("State-Level Demand Conversion", "How well is online demand converting into clinic revenue?")

    ecom_state = ecom["state_df"].head(12).copy()
    clinic_state_agg = clinic_ft.groupby("State").agg({"Qty": "sum", "Revenue": "sum", "Pincode": "nunique"}).reset_index()
    clinic_state_agg.columns = ["State", "Clinic_Qty", "Clinic_Rev", "Clinic_Pincodes"]
    clinic_state_agg["Clinic_Rev_Cr"] = clinic_state_agg["Clinic_Rev"] / 1e7

    state_comp = ecom_state.merge(clinic_state_agg, on="State", how="left")
    state_comp["Clinic_Rev_Cr"] = state_comp["Clinic_Rev_Cr"].fillna(0)
    state_comp["Multiplier"] = state_comp.apply(
        lambda r: round(r["Clinic_Rev_Cr"] / r["Rev_Cr"], 1) if r["Clinic_Rev_Cr"] > 0 and r["Rev_Cr"] > 0 else 0, axis=1
    )
    state_comp["Signal"] = state_comp["Multiplier"].apply(
        lambda x: "Strong Conversion" if x >= 2.0 else "Moderate" if x >= 1.0 else "Under-penetrated" if x > 0 else "No Clinic Presence"
    )

    st.dataframe(
        state_comp[["State", "Rev_Cr", "Clinic_Rev_Cr", "Multiplier", "Signal"]].rename(columns={
            "Rev_Cr": "E-Com Rev (Cr)", "Clinic_Rev_Cr": "Clinic Rev (Cr)"
        }),
        use_container_width=True, hide_index=True,
        column_config={
            "Multiplier": st.column_config.NumberColumn("Multiplier", format="%.1fx"),
            "Signal": st.column_config.TextColumn("Signal"),
        }
    )


# ================================================================
# TAB 4: ZONE INTELLIGENCE
# ================================================================
with tab4:
    # Zone Revenue vs Show Rate
    section_header("Zone Revenue & Conversion Efficiency", "Revenue (bars) vs Show Rate % (line)")

    fig_zone = make_subplots(specs=[[{"secondary_y": True}]])
    for _, row in zdf.iterrows():
        z = row["Zone"]
        color = ZONE_COLORS.get(z, "#666")
        show_info = zone_show_map.get(z, {})
        fig_zone.add_trace(
            go.Bar(x=[z], y=[row["Rev_Cr"]], name=z, marker_color=color, opacity=0.75,
                   showlegend=False, hovertemplate=f"{z}<br>Rs.{row['Rev_Cr']}Cr<extra></extra>"),
            secondary_y=False,
        )

    show_rates = []
    zone_labels = []
    for _, row in zdf.iterrows():
        z = row["Zone"]
        si = zone_show_map.get(z, {})
        show_rates.append(si.get("avg_show_pct", 0))
        zone_labels.append(z)

    fig_zone.add_trace(
        go.Scatter(x=zone_labels, y=show_rates, name="Show Rate %",
                   line=dict(color="#fff", width=2.5),
                   mode="lines+markers", marker=dict(size=8, color="#fff", line=dict(color="#0A0A0F", width=2))),
        secondary_y=True,
    )
    fig_zone.update_layout(**PLOTLY_LAYOUT, height=340, title="")
    fig_zone.update_yaxes(title_text="Revenue (Rs. Cr)", secondary_y=False, gridcolor="rgba(255,255,255,0.04)")
    fig_zone.update_yaxes(title_text="Show Rate %", secondary_y=True, range=[10, 30], gridcolor="rgba(255,255,255,0.02)")
    st.plotly_chart(fig_zone, use_container_width=True)

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    # Zone Deep Cards
    section_header("Zone Deep Dive Cards", "Revenue share, top clinic, and key metrics per zone")

    clinic_by_zone = clinic_ntb["clinic_df"].copy()

    zone_cols = st.columns(2)
    for i, (_, zrow) in enumerate(zdf.iterrows()):
        z = zrow["Zone"]
        color = ZONE_COLORS.get(z, "#666")
        show_info = zone_show_map.get(z, {})
        show_pct = show_info.get("avg_show_pct", 0)
        total_appt = show_info.get("total_appt", 0)
        rev_share = (zrow["Rev_Cr"] / zdf["Rev_Cr"].sum() * 100) if zdf["Rev_Cr"].sum() > 0 else 0
        zone_clinics = clinic_by_zone[clinic_by_zone["Zone"] == z].sort_values("Revenue", ascending=False)
        top_clinic = zone_clinics.iloc[0] if len(zone_clinics) > 0 else None
        n_clinics = len(zone_clinics)
        show_color = "#2EC4B6" if show_pct >= 25 else "#F0C808" if show_pct >= 20 else "#F15BB5"

        top_clinic_html = ""
        if top_clinic is not None:
            top_clinic_html = f"""
            <div style="background:rgba(255,255,255,0.02);border-radius:8px;padding:10px;border:1px solid rgba(255,255,255,0.04);margin-top:12px;">
                <div style="font-size:10px;color:#555;margin-bottom:3px;">Top Performer</div>
                <div style="display:flex;justify-content:space-between;align-items:center;">
                    <span style="font-size:13px;font-weight:700;color:#fff;">{top_clinic['Clinic']}</span>
                    <span style="font-size:12px;color:{color};font-weight:600;">Rs.{top_clinic['Rev_L']:.0f}L</span>
                </div>
            </div>
            """

        with zone_cols[i % 2]:
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,{color}08,transparent);border:1px solid {color}20;border-radius:16px;padding:22px;margin-bottom:16px;">
                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;">
                    <div>
                        <div style="font-size:20px;font-weight:800;color:{color};">{z}</div>
                        <span style="font-size:11px;color:#555;">{n_clinics} clinics</span>
                    </div>
                    <div style="text-align:right;">
                        <div style="font-size:22px;font-weight:700;color:#fff;">Rs.{zrow['Rev_Cr']}Cr</div>
                        <div style="font-size:11px;color:#666;">{rev_share:.0f}% share</div>
                    </div>
                </div>
                <div style="display:flex;gap:16px;">
                    <div style="flex:1;background:rgba(255,255,255,0.03);border-radius:8px;padding:10px;text-align:center;">
                        <div style="font-size:10px;color:#555;">NTB Qty</div>
                        <div style="font-size:14px;font-weight:700;color:#fff;">{zrow['Qty']:,}</div>
                    </div>
                    <div style="flex:1;background:rgba(255,255,255,0.03);border-radius:8px;padding:10px;text-align:center;">
                        <div style="font-size:10px;color:#555;">Appointments</div>
                        <div style="font-size:14px;font-weight:700;color:#fff;">{total_appt:,}</div>
                    </div>
                    <div style="flex:1;background:rgba(255,255,255,0.03);border-radius:8px;padding:10px;text-align:center;">
                        <div style="font-size:10px;color:#555;">Show %</div>
                        <div style="font-size:14px;font-weight:700;color:{show_color};">{show_pct:.1f}%</div>
                    </div>
                </div>
                {top_clinic_html}
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    # Monthly Visits Trajectory
    section_header("Monthly Visits Trajectory", "Tracking the growth flywheel across all clinics")

    mdf_visits = clinic_ntb["month_df"].copy()
    fig_visits = go.Figure()
    fig_visits.add_trace(go.Scatter(
        x=mdf_visits["month"], y=mdf_visits["visits"], name="Clinic Visits",
        fill="tozeroy", fillcolor="rgba(155,93,229,0.08)",
        line=dict(color="#9B5DE5", width=2.5), mode="lines+markers",
        marker=dict(size=4, color="#9B5DE5"),
    ))
    fig_visits.update_layout(**PLOTLY_LAYOUT, height=300, title="")
    fig_visits.update_yaxes(title_text="Total Visits")
    st.plotly_chart(fig_visits, use_container_width=True)

    # Zone Monthly Trend
    if not clinic_ntb["zone_month_df"].empty:
        section_header("Zone Revenue Trends", "Monthly revenue trajectory by zone")

        zmdf = clinic_ntb["zone_month_df"].copy()
        zmdf["rev_lakhs"] = zmdf["rev"] / 1e5
        fig_zone_trend = go.Figure()
        for z in zmdf["Zone"].unique():
            if z in ZONE_COLORS:
                zdata = zmdf[zmdf["Zone"] == z].sort_values("month")
                fig_zone_trend.add_trace(go.Scatter(
                    x=zdata["month"], y=zdata["rev_lakhs"], name=z,
                    line=dict(color=ZONE_COLORS[z], width=2), mode="lines",
                ))
        fig_zone_trend.update_layout(**PLOTLY_LAYOUT, height=320, title="")
        fig_zone_trend.update_yaxes(title_text="Revenue (Rs. Lakhs)")
        st.plotly_chart(fig_zone_trend, use_container_width=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# FOOTER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<div style="margin-top:40px;padding:16px 0;border-top:1px solid rgba(255,255,255,0.04);display:flex;justify-content:space-between;font-size:11px;color:#333;">
    <span>Gynoveda Expansion Intelligence &middot; Data: Jan 2025 - Jan 2026</span>
    <span>61 Clinics &middot; 7,415 Pincodes &middot; 6 Zones</span>
</div>
""", unsafe_allow_html=True)
